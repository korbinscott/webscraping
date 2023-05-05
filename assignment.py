from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
import keys
from twilio.rest import Client


url = 'https://www.coingecko.com/'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(url, headers=headers)

page = urlopen(req).read()

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)

wb = xl.Workbook()

ws = wb.active

ws.title = 'Bitcoin Report'

ws['A1'] = 'Name'
ws['B1'] = 'Ticker'
ws['C1'] = 'Current Price'
ws['D1'] = '% Change'
ws['E1'] = 'New Price'

# print(table_row[1])
# print(" ".join(coin_name.text.split()))

table_row = soup.findAll("tr")

for x in range(1,6):
    td = table_row[x].findAll('span')
    name = td[0].text
    ticker = td[1].text
    price = float(td[2].text.replace(",","").replace("$",""))
    change = td[4].text
    
    if change != 0:
        new_price = price * ((float(change.replace("%","")) * .001) +1)
    else:
        new_price = price
    
    ws['A' + str(x+1)] = name
    ws['B' + str(x+1)] = ticker
    ws['C' + str(x+1)] = price
    ws['D' + str(x+1)] = change
    ws['E' + str(x+1)] = new_price

ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10

header_font = Font(bold=True)

for cell in ws[1:1]:
    cell.font = header_font

wb.save("BitcoinReport.xlsx")


client = Client(keys.accountSID, keys.auth_token)

TwilioNumber = "+18449703428"

mycellphone = "+18066203022"

for x in range(1,6):
    td = table_row[x].findAll('span')
    name = td[0].text
    change = float(td[4].text.replace("%",""))
    price = float(td[2].text.replace(",","").replace("$",""))
    change = float(td[4].text.replace("%",""))

    if change != 0:
        new_price = price * ((change * .001) +1)
    else:
        new_price = price

    if new_price - price > 5:
        textmessage = client.messages.create(to=mycellphone,from_=TwilioNumber,body=f"The price for {name} has gone up!")
    elif new_price - price < -5:
        textmessage = client.messages.create(to=mycellphone,from_=TwilioNumber,body=f"The price for {name} has gone down!")

